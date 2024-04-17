from openpyxl import load_workbook
import os

# 현재 실행 파일 경로 가져오고, 입력파일 지정하기
program_path = os.path.abspath(__file__)
path = os.path.dirname(program_path)
input_file = path + '/input/' + input('입력 파일 : ')
# 입력 예시 : sales_2013.xlsx

# 엑셀 통합 문서 열기
workbook = load_workbook(input_file)
print('엑셀 통합 문서의 워크시트 수 : ', len(workbook.sheetnames))
for worksheet in workbook:
    print('워크시트 이름 : ', worksheet.title )
    print('워크시트 최대 행 수 : ', worksheet.max_row )
    print('워크시트 최대 열 수 : ', worksheet.max_column )
